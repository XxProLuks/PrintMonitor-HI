"""
Módulo para exportação avançada de dados (CSV, Excel, PNG, PDF) - Versão Profissional Hospitalar
"""
import sqlite3
import pandas as pd
import io
import base64
from typing import Dict, Optional, List, Tuple
from datetime import datetime
import logging

logger = logging.getLogger(__name__)

try:
    from reportlab.lib.pagesizes import letter, A4
    from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer, Image
    from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
    from reportlab.lib.units import inch
    from reportlab.lib import colors
    REPORTLAB_AVAILABLE = True
except ImportError:
    REPORTLAB_AVAILABLE = False
    logger.warning("reportlab não disponível. Exportação PDF limitada.")

try:
    from openpyxl import load_workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.chart import BarChart, PieChart, LineChart, Reference
    from openpyxl.utils import get_column_letter
    OPENPYXL_AVAILABLE = True
except ImportError:
    OPENPYXL_AVAILABLE = False
    logger.warning("openpyxl não disponível. Formatação Excel limitada.")


def exportar_csv(conn: sqlite3.Connection, query: str, params: tuple = ()) -> io.BytesIO:
    """Exporta dados para CSV"""
    try:
        df = pd.read_sql_query(query, conn, params=params)
        output = io.BytesIO()
        df.to_csv(output, index=False, encoding='utf-8-sig')
        output.seek(0)
        return output
    except Exception as e:
        logger.error(f"Erro ao exportar CSV: {e}")
        raise


def formatar_planilha_excel(worksheet, titulo: str = None):
    """Aplica formatação profissional a uma planilha Excel"""
    if not OPENPYXL_AVAILABLE:
        return
    
    try:
        # Cores profissionais hospitalares
        cor_cabecalho = "0066CC"  # Azul hospitalar
        cor_linha_par = "F5F5F5"  # Cinza claro
        cor_borda = "CCCCCC"
        
        # Estilo de cabeçalho
        header_fill = PatternFill(start_color=cor_cabecalho, end_color=cor_cabecalho, fill_type="solid")
        header_font = Font(bold=True, color="FFFFFF", size=11)
        header_alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        
        # Estilo de borda
        thin_border = Border(
            left=Side(style='thin', color=cor_borda),
            right=Side(style='thin', color=cor_borda),
            top=Side(style='thin', color=cor_borda),
            bottom=Side(style='thin', color=cor_borda)
        )
        
        # Aplica formatação ao cabeçalho (primeira linha)
        if worksheet.max_row > 0:
            for cell in worksheet[1]:
                cell.fill = header_fill
                cell.font = header_font
                cell.alignment = header_alignment
                cell.border = thin_border
        
        # Aplica formatação às linhas de dados
        for row_idx, row in enumerate(worksheet.iter_rows(min_row=2, max_row=worksheet.max_row), start=2):
            for cell in row:
                cell.border = thin_border
                cell.alignment = Alignment(horizontal="center", vertical="center")
                # Linhas alternadas
                if row_idx % 2 == 0:
                    cell.fill = PatternFill(start_color=cor_linha_par, end_color=cor_linha_par, fill_type="solid")
        
        # Ajusta largura das colunas
        for column in worksheet.columns:
            max_length = 0
            column_letter = get_column_letter(column[0].column)
            for cell in column:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            adjusted_width = min(max_length + 2, 50)
            worksheet.column_dimensions[column_letter].width = adjusted_width
        
        # Adiciona título se fornecido
        if titulo:
            worksheet.insert_rows(1)
            worksheet.merge_cells(f'A1:{get_column_letter(worksheet.max_column)}1')
            title_cell = worksheet['A1']
            title_cell.value = titulo
            title_cell.font = Font(bold=True, size=14, color="0066CC")
            title_cell.alignment = Alignment(horizontal="center", vertical="center")
            worksheet.row_dimensions[1].height = 25
            
            # Ajusta cabeçalho para linha 2
            if worksheet.max_row > 1:
                for cell in worksheet[2]:
                    cell.fill = header_fill
                    cell.font = header_font
                    cell.alignment = header_alignment
                    cell.border = thin_border
        
        # Congela primeira linha (cabeçalho)
        worksheet.freeze_panes = 'A2' if not titulo else 'A3'
        
    except Exception as e:
        logger.warning(f"Erro ao formatar planilha: {e}")


def exportar_excel(conn: sqlite3.Connection, query: str, params: tuple = ()) -> io.BytesIO:
    """Exporta dados para Excel com formatação básica"""
    try:
        df = pd.read_sql_query(query, conn, params=params)
        output = io.BytesIO()
        with pd.ExcelWriter(output, engine='openpyxl') as writer:
            df.to_excel(writer, index=False, sheet_name='Dados')
            
            # Aplica formatação se openpyxl estiver disponível
            if OPENPYXL_AVAILABLE:
                worksheet = writer.sheets['Dados']
                formatar_planilha_excel(worksheet)
        
        output.seek(0)
        return output
    except Exception as e:
        logger.error(f"Erro ao exportar Excel: {e}")
        raise


def exportar_relatorio_excel_hospitalar(
    conn: sqlite3.Connection,
    hospital_nome: str = "Hospital",
    start_date: Optional[str] = None,
    end_date: Optional[str] = None
) -> io.BytesIO:
    """
    Gera relatório Excel completo e profissional para ambiente hospitalar
    
    Args:
        conn: Conexão com banco de dados
        hospital_nome: Nome do hospital
        start_date: Data inicial (formato YYYY-MM-DD)
        end_date: Data final (formato YYYY-MM-DD)
    
    Returns:
        BytesIO: Buffer com o Excel gerado
    """
    output = io.BytesIO()
    
    # Constrói cláusula WHERE
    where_clause = "WHERE 1=1"
    params = []
    if start_date:
        where_clause += " AND date(date) >= date(?)"
        params.append(start_date)
    if end_date:
        where_clause += " AND date(date) <= date(?)"
        params.append(end_date)
    params_tuple = tuple(params) if params else ()
    
    with pd.ExcelWriter(output, engine='openpyxl') as writer:
        # ====================================================================
        # PLANILHA 1: SUMÁRIO EXECUTIVO
        # ====================================================================
        cursor = conn.cursor()
        # Usa módulo centralizado de cálculos
        from modules.calculo_impressao import calcular_folhas_fisicas
        
        # Verifica se job_id existe
        existing_columns = [col[1] for col in cursor.execute("PRAGMA table_info(events)").fetchall()]
        has_job_id = 'job_id' in existing_columns
        
        # Função auxiliar para obter cláusula GROUP BY de job
        def get_job_group_by(alias=''):
            prefix = f"{alias}." if alias else ""
            if has_job_id:
                return f"""CASE 
                    WHEN {prefix}job_id IS NOT NULL AND {prefix}job_id != '' THEN 
                        {prefix}job_id || '|' || COALESCE({prefix}printer_name, '') || '|' || {prefix}date
                    ELSE 
                        {prefix}user || '|' || {prefix}machine || '|' || COALESCE({prefix}document, '') || '|' || COALESCE({prefix}printer_name, '') || '|' || {prefix}date
                END"""
            else:
                return f"""{prefix}user || '|' || {prefix}machine || '|' || COALESCE({prefix}document, '') || '|' || COALESCE({prefix}printer_name, '') || '|' || {prefix}date"""
        
        # Total de impressões (jobs únicos, não eventos)
        if has_job_id:
            total_impressoes = cursor.execute(
                f"""SELECT COUNT(DISTINCT CASE 
                    WHEN job_id IS NOT NULL AND job_id != '' THEN 
                        job_id || '|' || COALESCE(printer_name, '') || '|' || date
                    ELSE 
                        user || '|' || machine || '|' || COALESCE(document, '') || '|' || COALESCE(printer_name, '') || '|' || date
                END) FROM events {where_clause}""",
                params_tuple
            ).fetchone()[0]
        else:
            total_impressoes = cursor.execute(
                f"""SELECT COUNT(DISTINCT user || '|' || machine || '|' || COALESCE(document, '') || '|' || COALESCE(printer_name, '') || '|' || date) 
                FROM events {where_clause}""",
                params_tuple
            ).fetchone()[0]
        
        # Total de páginas (folhas físicas) - AGRUPA POR JOB PRIMEIRO
        if has_job_id:
            rows = cursor.execute(
                f"""SELECT 
                    MAX(pages_printed) as pages,
                    MAX(COALESCE(duplex, 0)) as duplex
                FROM events {where_clause}
                GROUP BY CASE 
                    WHEN job_id IS NOT NULL AND job_id != '' THEN 
                        job_id || '|' || COALESCE(printer_name, '') || '|' || date
                    ELSE 
                        user || '|' || machine || '|' || COALESCE(document, '') || '|' || COALESCE(printer_name, '') || '|' || date
                END""",
                params_tuple
            ).fetchall()
        else:
            rows = cursor.execute(
                f"""SELECT 
                    MAX(pages_printed) as pages,
                    MAX(COALESCE(duplex, 0)) as duplex
                FROM events {where_clause}
                GROUP BY user || '|' || machine || '|' || COALESCE(document, '') || '|' || COALESCE(printer_name, '') || '|' || date""",
                params_tuple
            ).fetchall()
        
        total_paginas = 0
        for row in rows:
            pages = row[0] or 0
            duplex = row[1] if len(row) > 1 else None
            total_paginas += calcular_folhas_fisicas(pages, duplex)
        
        # Outras estatísticas
        total_usuarios = cursor.execute(
            f"SELECT COUNT(DISTINCT user) FROM events {where_clause}",
            params_tuple
        ).fetchone()[0]
        
        total_impressoras = cursor.execute(
            f"SELECT COUNT(DISTINCT printer_name) FROM events {where_clause} AND printer_name IS NOT NULL AND printer_name != ''",
            params_tuple
        ).fetchone()[0]
        
        # Custo total (soma de todos os eventos, não agrupado por job)
        custo_total = cursor.execute(
            f"SELECT SUM(CASE WHEN cost IS NOT NULL THEN cost ELSE 0 END) FROM events {where_clause}",
            params_tuple
        ).fetchone()[0] or 0
        
        # Média de páginas por impressão
        media_paginas = total_paginas / total_impressoes if total_impressoes > 0 else 0
        
        stats = (total_impressoes, total_paginas, total_usuarios, total_impressoras, custo_total, media_paginas)
        
        resumo_data = {
            'Métrica': [
                'Total de Impressões',
                'Total de Páginas',
                'Usuários Únicos',
                'Impressoras Monitoradas',
                'Custo Total (R$)',
                'Média de Páginas por Impressão',
                'Custo Médio por Página (R$)'
            ],
            'Valor': [
                stats[0] or 0,
                stats[1] or 0,
                stats[2] or 0,
                stats[3] or 0,
                f"R$ {stats[4]:,.2f}" if stats[4] else "R$ 0.00",
                f"{stats[5]:.2f}" if stats[5] else "0.00",
                f"R$ {stats[4]/stats[1]:.4f}" if stats[1] and stats[1] > 0 and stats[4] else "R$ 0.00"
            ]
        }
        
        df_resumo = pd.DataFrame(resumo_data)
        df_resumo.to_excel(writer, index=False, sheet_name='Sumário Executivo')
        
        if OPENPYXL_AVAILABLE:
            ws = writer.sheets['Sumário Executivo']
            formatar_planilha_excel(ws, f"{hospital_nome} - Sumário Executivo")
        
        # ====================================================================
        # PLANILHA 2: ANÁLISE POR SETOR - AGRUPA POR JOB PRIMEIRO
        # ====================================================================
        # Busca jobs únicos agrupados por job primeiro
        if has_job_id:
            setor_rows = cursor.execute(
                f"""SELECT 
                    MAX(COALESCE(account, 'Não especificado')) as setor,
                    MAX(pages_printed) as pages,
                    MAX(COALESCE(duplex, 0)) as duplex,
                    MAX(user) as user
                FROM events
                {where_clause}
                GROUP BY {get_job_group_by()}
                """, params_tuple
            ).fetchall()
        else:
            setor_rows = cursor.execute(
                f"""SELECT 
                    MAX(COALESCE(account, 'Não especificado')) as setor,
                    MAX(pages_printed) as pages,
                    MAX(COALESCE(duplex, 0)) as duplex,
                    MAX(user) as user
                FROM events
                {where_clause}
                GROUP BY {get_job_group_by()}
                """, params_tuple
            ).fetchall()
        
        # Agrupa por setor e calcula totais
        setores_dict = {}
        for row in setor_rows:
            setor = row[0]
            pages = row[1] or 0
            duplex = row[2] if len(row) > 2 else None
            user = row[3] if len(row) > 3 else None
            folhas = calcular_folhas_fisicas(pages, duplex)
            
            if setor not in setores_dict:
                setores_dict[setor] = {"impressoes": 0, "paginas": 0, "usuarios": set(), "total_pages": 0}
            setores_dict[setor]["impressoes"] += 1
            setores_dict[setor]["paginas"] += folhas
            setores_dict[setor]["total_pages"] += pages
            if user:
                setores_dict[setor]["usuarios"].add(user)
        
        # Calcula custo (soma de todos os eventos)
        custo_por_setor = {}
        custo_rows = cursor.execute(
            f"""SELECT 
                COALESCE(account, 'Não especificado') as setor,
                SUM(CASE WHEN cost IS NOT NULL THEN cost ELSE 0 END) as custo
            FROM events
            {where_clause}
            GROUP BY account
            """, params_tuple
        ).fetchall()
        for row in custo_rows:
            custo_por_setor[row[0]] = row[1] or 0
        
        setores_data = [
            (setor, data["impressoes"], data["paginas"], len(data["usuarios"]), 
             custo_por_setor.get(setor, 0), data["total_pages"] / data["impressoes"] if data["impressoes"] > 0 else 0)
            for setor, data in sorted(setores_dict.items(), key=lambda x: x[1]["paginas"], reverse=True)
        ]
        
        setores_df = pd.DataFrame(setores_data, columns=[
            'Setor/Departamento', 'Total Impressões', 'Total Páginas',
            'Usuários', 'Custo (R$)', 'Média Páginas/Job'
        ])
        setores_df.to_excel(writer, index=False, sheet_name='Análise por Setor')
        
        if OPENPYXL_AVAILABLE:
            ws = writer.sheets['Análise por Setor']
            formatar_planilha_excel(ws, f"{hospital_nome} - Análise por Setor/Departamento")
        
        # ====================================================================
        # PLANILHA 3: TOP USUÁRIOS - AGRUPA POR JOB PRIMEIRO
        # ====================================================================
        if has_job_id:
            usuario_rows = cursor.execute(
                f"""SELECT 
                    MAX(user) as user,
                    MAX(pages_printed) as pages,
                    MAX(COALESCE(duplex, 0)) as duplex,
                    MAX(date) as date
                FROM events
                {where_clause}
                GROUP BY {get_job_group_by()}, user
                """, params_tuple
            ).fetchall()
        else:
            usuario_rows = cursor.execute(
                f"""SELECT 
                    MAX(user) as user,
                    MAX(pages_printed) as pages,
                    MAX(COALESCE(duplex, 0)) as duplex,
                    MAX(date) as date
                FROM events
                {where_clause}
                GROUP BY {get_job_group_by()}, user
                """, params_tuple
            ).fetchall()
        
        # Agrupa por usuário e calcula totais
        usuarios_dict = {}
        for row in usuario_rows:
            user = row[0]
            pages = row[1] or 0
            duplex = row[2] if len(row) > 2 else None
            date_val = row[3] if len(row) > 3 else None
            folhas = calcular_folhas_fisicas(pages, duplex)
            
            if user not in usuarios_dict:
                usuarios_dict[user] = {"impressoes": 0, "paginas": 0, "total_pages": 0, "ultima_data": date_val}
            usuarios_dict[user]["impressoes"] += 1
            usuarios_dict[user]["paginas"] += folhas
            usuarios_dict[user]["total_pages"] += pages
            if date_val and (not usuarios_dict[user]["ultima_data"] or date_val > usuarios_dict[user]["ultima_data"]):
                usuarios_dict[user]["ultima_data"] = date_val
        
        # Calcula custo
        custo_por_usuario = {}
        custo_rows = cursor.execute(
            f"""SELECT 
                user,
                SUM(CASE WHEN cost IS NOT NULL THEN cost ELSE 0 END) as custo
            FROM events
            {where_clause}
            GROUP BY user
            """, params_tuple
        ).fetchall()
        for row in custo_rows:
            custo_por_usuario[row[0]] = row[1] or 0
        
        usuarios_data = [
            (user, data["impressoes"], data["paginas"], 
             data["total_pages"] / data["impressoes"] if data["impressoes"] > 0 else 0,
             custo_por_usuario.get(user, 0), data["ultima_data"])
            for user, data in sorted(usuarios_dict.items(), key=lambda x: x[1]["paginas"], reverse=True)[:50]
        ]
        
        usuarios_df = pd.DataFrame(usuarios_data, columns=[
            'Usuário', 'Total Impressões', 'Total Páginas',
            'Média Páginas/Job', 'Custo (R$)', 'Última Impressão'
        ])
        usuarios_df.to_excel(writer, index=False, sheet_name='Top Usuários')
        
        if OPENPYXL_AVAILABLE:
            ws = writer.sheets['Top Usuários']
            formatar_planilha_excel(ws, f"{hospital_nome} - Top 50 Usuários")
        
        # ====================================================================
        # PLANILHA 4: ANÁLISE DE IMPRESSORAS - AGRUPA POR JOB PRIMEIRO
        # ====================================================================
        # Verifica se tabela impressoras existe
        cursor.execute("SELECT name FROM sqlite_master WHERE type='table' AND name='impressoras'")
        has_impressoras_table = cursor.fetchone() is not None
        
        if has_job_id:
            if has_impressoras_table:
                impressora_rows = cursor.execute(
                    f"""SELECT 
                        MAX(COALESCE(e.printer_name, 'Não especificado')) as impressora,
                        MAX(COALESCE(i.ip, '')) as ip,
                        MAX(e.pages_printed) as pages,
                        MAX(COALESCE(e.duplex, 0)) as duplex,
                        MAX(e.user) as user
                    FROM events e
                    LEFT JOIN impressoras i ON e.printer_name = i.nome
                    WHERE {where_clause.replace("WHERE ", "").replace("date(date)", "date(e.date)")} AND e.printer_name IS NOT NULL
                    GROUP BY {get_job_group_by('e')}, e.printer_name
                    """, params_tuple
                ).fetchall()
            else:
                impressora_rows = cursor.execute(
                    f"""SELECT 
                        MAX(COALESCE(printer_name, 'Não especificado')) as impressora,
                        '' as ip,
                        MAX(pages_printed) as pages,
                        MAX(COALESCE(duplex, 0)) as duplex,
                        MAX(user) as user
                    FROM events
                    {where_clause} AND printer_name IS NOT NULL
                    GROUP BY {get_job_group_by()}, printer_name
                    """, params_tuple
                ).fetchall()
        else:
            if has_impressoras_table:
                impressora_rows = cursor.execute(
                    f"""SELECT 
                        MAX(COALESCE(e.printer_name, 'Não especificado')) as impressora,
                        MAX(COALESCE(i.ip, '')) as ip,
                        MAX(e.pages_printed) as pages,
                        MAX(COALESCE(e.duplex, 0)) as duplex,
                        MAX(e.user) as user
                    FROM events e
                    LEFT JOIN impressoras i ON e.printer_name = i.nome
                    WHERE {where_clause.replace("WHERE ", "").replace("date(date)", "date(e.date)")} AND e.printer_name IS NOT NULL
                    GROUP BY {get_job_group_by('e')}, e.printer_name
                    """, params_tuple
                ).fetchall()
            else:
                impressora_rows = cursor.execute(
                    f"""SELECT 
                        MAX(COALESCE(printer_name, 'Não especificado')) as impressora,
                        '' as ip,
                        MAX(pages_printed) as pages,
                        MAX(COALESCE(duplex, 0)) as duplex,
                        MAX(user) as user
                    FROM events
                    {where_clause} AND printer_name IS NOT NULL
                    GROUP BY {get_job_group_by()}, printer_name
                    """, params_tuple
                ).fetchall()
        
        # Agrupa por impressora e calcula totais
        impressoras_dict = {}
        for row in impressora_rows:
            impressora = row[0]
            ip = row[1] if len(row) > 1 else ''
            pages = row[2] if len(row) > 2 else 0
            duplex = row[3] if len(row) > 3 else None
            user = row[4] if len(row) > 4 else None
            folhas = calcular_folhas_fisicas(pages, duplex)
            
            if impressora not in impressoras_dict:
                impressoras_dict[impressora] = {
                    "ip": ip, "impressoes": 0, "paginas": 0, 
                    "total_pages": 0, "usuarios": set()
                }
            # Atualiza IP se encontrado
            if ip and not impressoras_dict[impressora]["ip"]:
                impressoras_dict[impressora]["ip"] = ip
            impressoras_dict[impressora]["impressoes"] += 1
            impressoras_dict[impressora]["paginas"] += folhas
            impressoras_dict[impressora]["total_pages"] += pages
            if user:
                impressoras_dict[impressora]["usuarios"].add(user)
        
        # Calcula custo
        custo_por_impressora = {}
        custo_rows = cursor.execute(
            f"""SELECT 
                printer_name,
                SUM(CASE WHEN cost IS NOT NULL THEN cost ELSE 0 END) as custo
            FROM events
            {where_clause} AND printer_name IS NOT NULL
            GROUP BY printer_name
            """, params_tuple
        ).fetchall()
        for row in custo_rows:
            custo_por_impressora[row[0]] = row[1] or 0
        
        impressoras_data = [
            (impressora, data.get("ip", "") or "", data["impressoes"], data["paginas"],
             data["total_pages"] / data["impressoes"] if data["impressoes"] > 0 else 0,
             custo_por_impressora.get(impressora, 0), len(data["usuarios"]))
            for impressora, data in sorted(impressoras_dict.items(), key=lambda x: x[1]["paginas"], reverse=True)[:30]
        ]
        
        impressoras_df = pd.DataFrame(impressoras_data, columns=[
            'Impressora', 'IP', 'Total Impressões', 'Total Páginas',
            'Média Páginas/Job', 'Custo (R$)', 'Usuários Únicos'
        ])
        impressoras_df.to_excel(writer, index=False, sheet_name='Análise de Impressoras')
        
        if OPENPYXL_AVAILABLE:
            ws = writer.sheets['Análise de Impressoras']
            formatar_planilha_excel(ws, f"{hospital_nome} - Análise de Impressoras")
        
        # ====================================================================
        # PLANILHA 5: MODO DE COR - AGRUPA POR JOB PRIMEIRO
        # ====================================================================
        if has_job_id:
            color_rows = cursor.execute(
                f"""SELECT 
                    MAX(color_mode) as color_mode,
                    MAX(pages_printed) as pages,
                    MAX(COALESCE(duplex, 0)) as duplex
                FROM events
                {where_clause} AND color_mode IS NOT NULL
                GROUP BY {get_job_group_by()}
                """, params_tuple
            ).fetchall()
        else:
            color_rows = cursor.execute(
                f"""SELECT 
                    MAX(color_mode) as color_mode,
                    MAX(pages_printed) as pages,
                    MAX(COALESCE(duplex, 0)) as duplex
                FROM events
                {where_clause} AND color_mode IS NOT NULL
                GROUP BY {get_job_group_by()}
                """, params_tuple
            ).fetchall()
        
        # Agrupa por color_mode e calcula totais
        color_dict = {}
        for row in color_rows:
            color_mode = row[0]
            if not color_mode:
                continue
            pages = row[1] or 0
            duplex = row[2] if len(row) > 2 else None
            folhas = calcular_folhas_fisicas(pages, duplex)
            
            if color_mode not in color_dict:
                color_dict[color_mode] = {"impressoes": 0, "paginas": 0}
            color_dict[color_mode]["impressoes"] += 1
            color_dict[color_mode]["paginas"] += folhas
        
        total_paginas_geral = sum(data["paginas"] for data in color_dict.values())
        color_data = [
            (color_mode or 'Não especificado', data["impressoes"], data["paginas"],
             round(data["paginas"] * 100.0 / total_paginas_geral, 2) if total_paginas_geral > 0 else 0)
            for color_mode, data in sorted(color_dict.items(), key=lambda x: x[1]["paginas"], reverse=True)
        ]
        
        color_df = pd.DataFrame(color_data, columns=[
            'Modo de Cor', 'Total Impressões', 'Total Páginas', '% do Total'
        ])
        color_df.to_excel(writer, index=False, sheet_name='Modo de Cor')
        
        if OPENPYXL_AVAILABLE:
            ws = writer.sheets['Modo de Cor']
            formatar_planilha_excel(ws, f"{hospital_nome} - Análise por Modo de Cor")
        
        # ====================================================================
        # PLANILHA 6: ANÁLISE DE DUPLEX - AGRUPA POR JOB PRIMEIRO
        # ====================================================================
        if has_job_id:
            duplex_rows = cursor.execute(
                f"""SELECT 
                    MAX(COALESCE(duplex, 0)) as duplex,
                    MAX(pages_printed) as pages
                FROM events
                {where_clause}
                GROUP BY {get_job_group_by()}
                """, params_tuple
            ).fetchall()
        else:
            duplex_rows = cursor.execute(
                f"""SELECT 
                    MAX(COALESCE(duplex, 0)) as duplex,
                    MAX(pages_printed) as pages
                FROM events
                {where_clause}
                GROUP BY {get_job_group_by()}
                """, params_tuple
            ).fetchall()
        
        # Agrupa por tipo duplex e calcula totais
        duplex_dict = {"Duplex (Economia)": {"impressoes": 0, "paginas": 0}, 
                       "Simples": {"impressoes": 0, "paginas": 0}, 
                       "Não especificado": {"impressoes": 0, "paginas": 0}}
        for row in duplex_rows:
            duplex_val = row[0]
            pages = row[1] or 0
            folhas = calcular_folhas_fisicas(pages, duplex_val)
            
            if duplex_val == 1:
                tipo = "Duplex (Economia)"
            elif duplex_val == 0:
                tipo = "Simples"
            else:
                tipo = "Não especificado"
            
            duplex_dict[tipo]["impressoes"] += 1
            duplex_dict[tipo]["paginas"] += folhas
        
        total_paginas_geral = sum(data["paginas"] for data in duplex_dict.values())
        duplex_data = [
            (tipo, data["impressoes"], data["paginas"],
             round(data["paginas"] * 100.0 / total_paginas_geral, 2) if total_paginas_geral > 0 else 0)
            for tipo, data in duplex_dict.items() if data["impressoes"] > 0
        ]
        
        duplex_df = pd.DataFrame(duplex_data, columns=[
            'Tipo', 'Total Impressões', 'Total Páginas', '% do Total'
        ])
        duplex_df.to_excel(writer, index=False, sheet_name='Análise Duplex')
        
        if OPENPYXL_AVAILABLE:
            ws = writer.sheets['Análise Duplex']
            formatar_planilha_excel(ws, f"{hospital_nome} - Análise de Uso Duplex")
        
        # ====================================================================
        # PLANILHA 7: ANÁLISE TEMPORAL (Por Dia da Semana) - AGRUPA POR JOB PRIMEIRO
        # ====================================================================
        if has_job_id:
            dias_rows = cursor.execute(
                f"""SELECT 
                    CASE CAST(strftime('%w', date) AS INTEGER)
                        WHEN 0 THEN 'Domingo'
                        WHEN 1 THEN 'Segunda-feira'
                        WHEN 2 THEN 'Terça-feira'
                        WHEN 3 THEN 'Quarta-feira'
                        WHEN 4 THEN 'Quinta-feira'
                        WHEN 5 THEN 'Sexta-feira'
                        WHEN 6 THEN 'Sábado'
                    END as dia_semana,
                    MAX(pages_printed) as pages,
                    MAX(COALESCE(duplex, 0)) as duplex,
                    strftime('%w', date) as order_val
                FROM events
                {where_clause}
                GROUP BY {get_job_group_by()}, strftime('%w', date)
                """, params_tuple
            ).fetchall()
        else:
            dias_rows = cursor.execute(
                f"""SELECT 
                    CASE CAST(strftime('%w', date) AS INTEGER)
                        WHEN 0 THEN 'Domingo'
                        WHEN 1 THEN 'Segunda-feira'
                        WHEN 2 THEN 'Terça-feira'
                        WHEN 3 THEN 'Quarta-feira'
                        WHEN 4 THEN 'Quinta-feira'
                        WHEN 5 THEN 'Sexta-feira'
                        WHEN 6 THEN 'Sábado'
                    END as dia_semana,
                    MAX(pages_printed) as pages,
                    MAX(COALESCE(duplex, 0)) as duplex,
                    strftime('%w', date) as order_val
                FROM events
                {where_clause}
                GROUP BY {get_job_group_by()}, strftime('%w', date)
                """, params_tuple
            ).fetchall()
        
        # Agrupa por dia da semana e calcula totais
        dias_dict = {}
        for row in dias_rows:
            dia = row[0]
            pages = row[1] or 0
            duplex = row[2] if len(row) > 2 else None
            order_val = int(row[3]) if len(row) > 3 and row[3] else 0
            folhas = calcular_folhas_fisicas(pages, duplex)
            
            if dia not in dias_dict:
                dias_dict[dia] = {"impressoes": 0, "paginas": 0, "total_pages": 0, "order": order_val}
            dias_dict[dia]["impressoes"] += 1
            dias_dict[dia]["paginas"] += folhas
            dias_dict[dia]["total_pages"] += pages
        
        # Ordena por dia da semana
        dias_data = [
            (dia, data["impressoes"], data["paginas"],
             data["total_pages"] / data["impressoes"] if data["impressoes"] > 0 else 0)
            for dia, data in sorted(dias_dict.items(), key=lambda x: x[1]["order"])
        ]
        
        dias_df = pd.DataFrame(dias_data, columns=[
            'Dia da Semana', 'Total Impressões', 'Total Páginas', 'Média Páginas/Job'
        ])
        dias_df.to_excel(writer, index=False, sheet_name='Análise Temporal')
        
        if OPENPYXL_AVAILABLE:
            ws = writer.sheets['Análise Temporal']
            formatar_planilha_excel(ws, f"{hospital_nome} - Análise por Dia da Semana")
        
        # ====================================================================
        # PLANILHA 8: DETALHAMENTO DE EVENTOS (Amostra)
        # ====================================================================
        # Verifica se tabela impressoras existe
        cursor.execute("SELECT name FROM sqlite_master WHERE type='table' AND name='impressoras'")
        has_impressoras_table = cursor.fetchone() is not None
        
        if has_impressoras_table:
            # Constrói WHERE com alias
            where_with_alias = where_clause.replace("WHERE ", "").replace("date(date)", "date(e.date)")
            cursor.execute(f"""
                SELECT 
                    e.date as 'Data/Hora',
                    e.user as 'Usuário',
                    COALESCE(e.printer_name, 'Não especificado') as 'Impressora',
                    COALESCE(i.ip, '') as 'IP',
                    e.pages_printed as 'Páginas',
                    COALESCE(e.color_mode, 'N/A') as 'Modo Cor',
                    CASE WHEN e.duplex = 1 THEN 'Sim' ELSE 'Não' END as 'Duplex',
                    COALESCE(e.document, 'N/A') as 'Documento',
                    CASE WHEN e.cost IS NOT NULL THEN e.cost ELSE 0 END as 'Custo (R$)'
                FROM events e
                LEFT JOIN impressoras i ON e.printer_name = i.nome
                WHERE {where_with_alias}
                ORDER BY e.date DESC
                LIMIT 1000
            """, params_tuple)
            
            eventos_df = pd.DataFrame(cursor.fetchall(), columns=[
                'Data/Hora', 'Usuário', 'Impressora', 'IP', 'Páginas',
                'Modo Cor', 'Duplex', 'Documento', 'Custo (R$)'
            ])
        else:
            cursor.execute(f"""
                SELECT 
                    date as 'Data/Hora',
                    user as 'Usuário',
                    COALESCE(printer_name, 'Não especificado') as 'Impressora',
                    '' as 'IP',
                    pages_printed as 'Páginas',
                    COALESCE(color_mode, 'N/A') as 'Modo Cor',
                    CASE WHEN duplex = 1 THEN 'Sim' ELSE 'Não' END as 'Duplex',
                    COALESCE(document, 'N/A') as 'Documento',
                    CASE WHEN cost IS NOT NULL THEN cost ELSE 0 END as 'Custo (R$)'
                FROM events
                {where_clause}
                ORDER BY date DESC
                LIMIT 1000
            """, params_tuple)
            
            eventos_df = pd.DataFrame(cursor.fetchall(), columns=[
                'Data/Hora', 'Usuário', 'Impressora', 'IP', 'Páginas',
                'Modo Cor', 'Duplex', 'Documento', 'Custo (R$)'
            ])
        eventos_df.to_excel(writer, index=False, sheet_name='Detalhamento (Amostra)')
        
        if OPENPYXL_AVAILABLE:
            ws = writer.sheets['Detalhamento (Amostra)']
            formatar_planilha_excel(ws, f"{hospital_nome} - Detalhamento de Eventos (Últimos 1000)")
        
        # ====================================================================
        # PLANILHA 9: METADADOS DO RELATÓRIO
        # ====================================================================
        metadata = {
            'Informação': [
                'Hospital/Instituição',
                'Data de Geração',
                'Período Analisado',
                'Total de Registros',
                'Versão do Sistema'
            ],
            'Valor': [
                hospital_nome,
                datetime.now().strftime('%d/%m/%Y %H:%M:%S'),
                f"{start_date or 'Início'} a {end_date or 'Fim'}",
                stats[0] or 0,
                'Print Monitor v3.0'
            ]
        }
        
        metadata_df = pd.DataFrame(metadata)
        metadata_df.to_excel(writer, index=False, sheet_name='Metadados')
        
        if OPENPYXL_AVAILABLE:
            ws = writer.sheets['Metadados']
            formatar_planilha_excel(ws, "Informações do Relatório")
    
    output.seek(0)
    return output


def exportar_grafico_png(chart_data: Dict) -> Optional[bytes]:
    """Exporta gráfico como PNG (requer matplotlib)"""
    try:
        import matplotlib
        matplotlib.use('Agg')
        import matplotlib.pyplot as plt
        
        fig, ax = plt.subplots(figsize=(10, 6))
        
        if chart_data.get('type') == 'bar':
            ax.bar(chart_data.get('labels', []), chart_data.get('values', []))
        elif chart_data.get('type') == 'line':
            ax.plot(chart_data.get('labels', []), chart_data.get('values', []))
        elif chart_data.get('type') == 'pie':
            ax.pie(chart_data.get('values', []), labels=chart_data.get('labels', []))
        
        ax.set_title(chart_data.get('title', 'Gráfico'))
        plt.tight_layout()
        
        output = io.BytesIO()
        plt.savefig(output, format='png', dpi=150, bbox_inches='tight')
        plt.close()
        output.seek(0)
        return output.read()
    except ImportError:
        logger.warning("matplotlib não disponível. Exportação PNG não suportada.")
        return None
    except Exception as e:
        logger.error(f"Erro ao exportar gráfico PNG: {e}")
        return None


def exportar_pdf_relatorio(conn: sqlite3.Connection, titulo: str, 
                          dados: Dict, graficos: Optional[list] = None) -> io.BytesIO:
    """Exporta relatório completo em PDF"""
    if not REPORTLAB_AVAILABLE:
        raise ImportError("reportlab não está instalado")
    
    try:
        output = io.BytesIO()
        doc = SimpleDocTemplate(output, pagesize=A4)
        story = []
        styles = getSampleStyleSheet()
        
        # Título
        title_style = ParagraphStyle(
            'CustomTitle',
            parent=styles['Heading1'],
            fontSize=24,
            textColor=colors.HexColor('#4a90e2'),
            spaceAfter=30,
            alignment=1  # Center
        )
        story.append(Paragraph(titulo, title_style))
        story.append(Spacer(1, 0.5*inch))
        
        # Dados em tabela
        if dados.get('table_data'):
            table_data = [dados['table_headers']] + dados['table_data']
            table = Table(table_data)
            table.setStyle(TableStyle([
                ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#4a90e2')),
                ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
                ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
                ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
                ('FONTSIZE', (0, 0), (-1, 0), 12),
                ('BOTTOMPADDING', (0, 0), (-1, 0), 12),
                ('BACKGROUND', (0, 1), (-1, -1), colors.beige),
                ('GRID', (0, 0), (-1, -1), 1, colors.black)
            ]))
            story.append(table)
            story.append(Spacer(1, 0.3*inch))
        
        # Gráficos
        if graficos:
            for grafico in graficos:
                if grafico.get('type') == 'image' and grafico.get('data'):
                    img_data = base64.b64decode(grafico['data'])
                    img = Image(io.BytesIO(img_data), width=6*inch, height=4*inch)
                    story.append(img)
                    story.append(Spacer(1, 0.2*inch))
        
        # Rodapé
        footer = Paragraph(
            f"Gerado em {datetime.now().strftime('%d/%m/%Y %H:%M:%S')}",
            styles['Normal']
        )
        story.append(Spacer(1, 0.5*inch))
        story.append(footer)
        
        doc.build(story)
        output.seek(0)
        return output
    except Exception as e:
        logger.error(f"Erro ao exportar PDF: {e}")
        raise
